"""
Usage:
    python restructure_to_excel.py <input_csv> [<input_csv2> ...]

Examples:
    python restructure_to_excel.py clip_pacs_results.csv
    python restructure_to_excel.py clip_pacs_results.csv clip_vlcs_results.csv
    python restructure_to_excel.py *.csv

Output:
    For each input CSV, an Excel file is saved in the same folder
    with the same name but a .xlsx extension.
    e.g. clip_pacs_results.csv  →  clip_pacs_results.xlsx
"""

import sys
import os
import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def style_cell(cell, font=None, fill=None, alignment=None, border=None, number_format=None):
    if font:          cell.font          = font
    if fill:          cell.fill          = fill
    if alignment:     cell.alignment     = alignment
    if border:        cell.border        = border
    if number_format: cell.number_format = number_format


def process_csv(csv_path: str):
    print(f"Processing: {csv_path}")

    df = pd.read_csv(csv_path)

    # ── Validate expected columns ────────────────────────────────────────
    required = {'train_domains', 'test_domains', 'fraction', 'method'}
    missing = required - set(df.columns)
    if missing:
        print(f"  ERROR: Missing columns {missing} — skipping.\n")
        return

    # Metric columns = everything that isn't a key column
    key_cols = ['train_domains', 'test_domains', 'fraction', 'method']
    metrics = [c for c in df.columns if c not in key_cols]
    method_cols = sorted(df['method'].unique(), key=lambda m: df[df['method']==m].index[0])

    # ── Styles ───────────────────────────────────────────────────────────
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    header_fill  = PatternFill("solid", fgColor="2F5496")
    subhead_font = Font(bold=True, size=10)
    subhead_fill = PatternFill("solid", fgColor="D9E1F2")
    frac_font    = Font(bold=True, size=10)
    frac_fill    = PatternFill("solid", fgColor="EEF2FF")
    center       = Alignment(horizontal="center", vertical="center")
    left         = Alignment(horizontal="left",   vertical="center")
    thin         = Side(style="thin", color="BFBFBF")
    border       = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ── Build workbook ───────────────────────────────────────────────────
    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    n_cols = 2 + len(method_cols)   # fraction + metric + one per method
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 14
    col_letters = [chr(ord('C') + i) for i in range(len(method_cols))]
    for letter in col_letters:
        ws.column_dimensions[letter].width = 13

    current_row = 1

    for (train_d, test_d), grp in df.groupby(['train_domains', 'test_domains'], sort=False):

        # ── Domain heading ───────────────────────────────────────────────
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row,   end_column=n_cols)
        cell = ws.cell(row=current_row, column=1,
                       value=f"Train: {train_d}    →    Test: {test_d}")
        style_cell(cell, font=header_font, fill=header_fill,
                   alignment=center, border=border)
        for c in range(2, n_cols + 1):
            style_cell(ws.cell(row=current_row, column=c), border=border)
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # ── Column sub-header ────────────────────────────────────────────
        for col_i, label in enumerate(['fraction', 'metric'] + list(method_cols), start=1):
            cell = ws.cell(row=current_row, column=col_i, value=label)
            style_cell(cell, font=subhead_font, fill=subhead_fill,
                       alignment=center, border=border)
        ws.row_dimensions[current_row].height = 16
        current_row += 1

        # ── Build NC pairs: maps each NC method to its base counterpart ────
        # e.g. "ERM+NC" → "ERM",  "CORAL+NC" → "CORAL"
        nc_pairs = {}
        for method in method_cols:
            if method.endswith('+NC'):
                base = method[:-3]   # strip "+NC"
                if base in method_cols:
                    nc_pairs[method] = base

        # ── Data rows ────────────────────────────────────────────────────
        fractions = sorted(grp['fraction'].unique())
        for f_idx, frac in enumerate(fractions):
            frac_grp = grp[grp['fraction'] == frac]

            for m_idx, metric in enumerate(metrics):
                frac_label = frac if m_idx == 0 else ""
                row_vals = [frac_label, metric]

                # Collect numeric values per method first (needed for comparison)
                method_values = {}
                for method in method_cols:
                    method_row = frac_grp[frac_grp['method'] == method]
                    method_values[method] = (
                        round(float(method_row[metric].values[0]), 4)
                        if not method_row.empty else None
                    )
                    row_vals.append(method_values[method] if method_values[method] is not None else "")

                for col_i, val in enumerate(row_vals, start=1):
                    cell = ws.cell(row=current_row, column=col_i, value=val)
                    if col_i == 1:
                        style_cell(cell, font=frac_font, fill=frac_fill,
                                   alignment=center, border=border)
                    elif col_i == 2:
                        style_cell(cell, alignment=left, border=border)
                    else:
                        # Check if this method's NC value beats its base
                        method = list(method_cols)[col_i - 3]
                        is_bold = False
                        if method in nc_pairs:
                            base = nc_pairs[method]
                            nc_val   = method_values.get(method)
                            base_val = method_values.get(base)
                            if nc_val is not None and base_val is not None:
                                # For losses: lower is better; for acc: higher is better
                                if 'loss' in metric:
                                    is_bold = nc_val < base_val
                                else:
                                    is_bold = nc_val > base_val
                        style_cell(cell,
                                   font=Font(bold=is_bold, size=10),
                                   alignment=center, border=border,
                                   number_format='0.0000')
                current_row += 1

            # Blank row between fractions (not after last)
            if f_idx < len(fractions) - 1:
                current_row += 1

        # Two blank rows between domain groups
        current_row += 2

    # ── Save ─────────────────────────────────────────────────────────────
    out_path = os.path.splitext(csv_path)[0] + ".xlsx"
    wb.save(out_path)
    print(f"  Saved → {out_path}\n")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    for csv_file in sys.argv[1:]:
        if not os.path.isfile(csv_file):
            print(f"  WARNING: File not found — {csv_file}\n")
            continue
        process_csv(csv_file)

    print("Done.")
